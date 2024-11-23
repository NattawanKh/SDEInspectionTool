import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from sqlite_dbcon import db_connect


def data_lot_file() :
        # Lot No Input
        global lot_no
        lot_no = 'TEST-2024'
        # Fetch device data from the database
        results = {
            # Device Data Part =========================================
            0: (mac_id := [], device_type := [], status := []),
            # Sumary Part ==============================================
            1: (totals := [], good_product := [], ng_product := []),
            # NG Product ===============================================
            2: (ng_mac_id := [], ng_device_type := [], issue_name := [])
            }
        #=================================================================================================================================
        f_select  = ["device_id, devices_type, inspec_note","qty_product, good_product, ng_product","device_id, devices_type, issue_name"]
        t_select  = ["db_sde.devices_income","db_sde.devices_income_lot","db_sde.devices_income"]
        condition = [f"lot_box_id = '{lot_no}'",f"lot_no = '{lot_no}'",f"inspec_note = 'NG' AND lot_box_id = '{lot_no}'"]
        rounds = [0,1,2]
        for round in rounds:
            this_con = db_connect()
            this_con.connect_select(t_select[round], condition[round], f_select[round])
            
            for data in this_con:
                for lst, value in zip(results[round], data):
                    lst.append(value)
        #=================================================================================================================================
        data2excel(mac_id,device_type,status,totals,good_product,ng_product,ng_mac_id,ng_device_type,issue_name)

def data2excel(mac_id,device_type,status,totals,good_product,ng_product,ng_mac_id,ng_device_type,issue_name) :
    global lot_no
    try:
        # Create a DataFrame for device data
        data = {'Device ID': mac_id, 'Device Type': device_type, 'Status': status}
        df = pd.DataFrame(data)

        # Save DataFrame to an Excel file with custom formatting
        file_path = 'device_data.xlsx'
        sheet_name=f'{lot_no}'
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Set font style and alignment for the entire sheet
            font = Font(name='SCG', size=12)
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Define border style
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = thin_border

            # Make headers bold and set background color
            header_font = Font(name='SCG', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for cell in worksheet[1]:  # Assuming the first row is the header
                cell.font = header_font
                cell.alignment = alignment
                cell.border = thin_border
                cell.fill = header_fill

            #===============================================================================================
            # Merge cells E2, F2, G2 and set text
            worksheet.merge_cells('E2:G2')
            worksheet['E2'] = 'Summary'
            worksheet['E2'].font = Font(name='SCG', size=12, bold=True, color='FFFFFF')
            worksheet['E2'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            worksheet['E2'].alignment = alignment
            worksheet['E2'].border = thin_border
            worksheet['F2'].border = thin_border
            worksheet['G2'].border = thin_border
            # Set text in cells E3, F3, G3
            worksheet['E3'] = 'Total'
            worksheet['F3'] = 'Good'
            worksheet['G3'] = 'NG'
            for cell in ['E3', 'F3', 'G3']:
                worksheet[cell].font = Font(name='SCG', size=12, bold=True)
                worksheet[cell].alignment = alignment
                worksheet[cell].border = thin_border
            # Insert total, good, and ng product values
            worksheet['E4'] = totals[0]
            worksheet['F4'] = good_product[0]
            worksheet['G4'] = ng_product[0]
            for cell in ['E4', 'F4', 'G4']:
                worksheet[cell].font = Font(name='SCG', size=12)
                worksheet[cell].alignment = alignment
                worksheet[cell].border = thin_border
            #===============================================================================================
            # Merge cells I2, J2, K2 and set text
            worksheet.merge_cells('I2:K2')
            worksheet['I2'] = 'NG Product'
            worksheet['I2'].font = Font(name='SCG', size=12, bold=True, color='FFFFFF')
            worksheet['I2'].fill = PatternFill(start_color="c00000", end_color="c00000", fill_type="solid")
            worksheet['I2'].alignment = alignment
            worksheet['I2'].border = thin_border
            worksheet['J2'].border = thin_border
            worksheet['K2'].border = thin_border
            # Set text in cells I3, J3, K3
            worksheet['I3'] = 'Device ID'
            worksheet['J3'] = 'Type'
            worksheet['K3'] = 'Defect'
            for cell in ['I3', 'J3', 'K3']:
                worksheet[cell].font = Font(name='SCG', size=12, bold=True,color='FFFFFF')
                worksheet[cell].fill = PatternFill(start_color="c00000", end_color="c00000", fill_type="solid")
                worksheet[cell].alignment = alignment
                worksheet[cell].border = thin_border
            # Insert NG product values
            for row, (mac_id, device_type, issue) in enumerate(zip(ng_mac_id, ng_device_type, issue_name), start=4):
                    worksheet.cell(row=row, column=9, value=mac_id).font = Font(name='SCG', size=12)
                    worksheet.cell(row=row, column=9, value=mac_id).fill = PatternFill(start_color="e6b8b7", end_color="e6b8b7", fill_type="solid")
                    worksheet.cell(row=row, column=9).alignment = alignment
                    worksheet.cell(row=row, column=9).border = thin_border
                    #====================================================================================================================================
                    worksheet.cell(row=row, column=10, value=device_type).font = Font(name='SCG', size=12)
                    worksheet.cell(row=row, column=10, value=device_type).fill = PatternFill(start_color="e6b8b7", end_color="e6b8b7", fill_type="solid")
                    worksheet.cell(row=row, column=10).alignment = alignment
                    worksheet.cell(row=row, column=10).border = thin_border
                    #====================================================================================================================================
                    worksheet.cell(row=row, column=11, value=issue).font = Font(name='SCG', size=12)
                    worksheet.cell(row=row, column=11, value=issue).fill = PatternFill(start_color="e6b8b7", end_color="e6b8b7", fill_type="solid")
                    worksheet.cell(row=row, column=11).alignment = alignment
                    worksheet.cell(row=row, column=11).border = thin_border
                    #====================================================================================================================================
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 6)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Color Pattern ===============================================================================
            # Apply conditional formatting for the "Status" column
            olive_green_fill = PatternFill(start_color="d8e4bc", end_color="d8e4bc", fill_type="solid")
            red_accent_fill = PatternFill(start_color="e6b8b7", end_color="e6b8b7", fill_type="solid")

            for cell in worksheet['C']:  # Assuming "Status" is in column C
                if cell.row == 1:
                    continue  # Skip header
                if cell.value == 'GOOD':
                    cell.fill = olive_green_fill
                elif cell.value == 'NG':
                    cell.fill = red_accent_fill
            #===============================================================================================
        print("Excel file created successfully.")

    except Exception as e:
        print(f"An error occurred: {e}")

data_lot_file()
